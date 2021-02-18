"""用于读写操作excel文件"""
import openpyxl
from core.logger import logger


class ReadExcel(object):
    # 读取excel数据的类
    def __init__(self, file_name):
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        :param sheet_name: 表单名 ———> str类型
        """
        # 打开文件
        self.wb = openpyxl.load_workbook(file_name)
        self.sheets = ["支出", "收入", "借入", "借出", "还款", "记录"]
        self.sheets_names = self.wb.sheetnames  # 获取excel文档中所有工作簿表

    def read_data(self):
        result = {}
        for sheet_name in self.sheets:
            logger.debug("reading %s" % sheet_name)
            if sheet_name not in self.sheets_names:
                # 不在自己需要的表明则跳出本次循环，不进行数据获取否则会出错
                continue
            # 选择表单
            self.sh = self.wb[sheet_name]
            # 按行读取数据转化为列表
            rows_data = list(self.sh.rows)
            # print(rows_data)
            # 获取表单的表头信息
            titles = []
            for title in rows_data[0]:
                titles.append(title.value)
            # print(titles)
            logger.info("读取到%s标题表头为：%s" % (sheet_name, titles))
            # 定义一个空列表用来存储测试用例
            cases = []
            for case in rows_data[1:]:
                # print(case)
                data = []
                for cell in case:
                    # 获取一条测试用例数据
                    data.append(cell.value)
                    # print(cell.value, type(cell.value))
                case_data = dict(list(zip(titles, data)))
                cases.append(case_data)
            result[sheet_name] = cases
            logger.debug("%s : %s" % (sheet_name, cases))
        return result


class WriteExcel(object):
    # 写出excel数据的类
    def __init__(self, file_name, result):
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        :param result: 数据 ———> dict类型  {"支出":[(,),(,)],"收入"}
        """
        # 打开文件
        self.wb = openpyxl.Workbook()
        self.result = result
        self.file_name = file_name

    def write_data(self):
        # 写出数据库所有数据到excel
        for sheet_name in self.result:
            self.sh = self.wb.create_sheet(sheet_name)
            # 设置列宽
            if sheet_name == "记录":
                self.sh.column_dimensions["A"].width = 13
                self.sh.column_dimensions["B"].width = 20
                self.sh.column_dimensions["C"].width = 30
                self.sh.column_dimensions["D"].width = 30
                self.sh.column_dimensions["E"].width = 35
                self.sh.column_dimensions["F"].width = 25
            else:
                self.sh.column_dimensions["A"].width = 13
                self.sh.column_dimensions["B"].width = 50
                self.sh.column_dimensions["C"].width = 15
                self.sh.column_dimensions["D"].width = 20
                self.sh.column_dimensions["E"].width = 40
                self.sh.column_dimensions["F"].width = 10
                self.sh.column_dimensions["G"].width = 10
                self.sh.column_dimensions["H"].width = 10
                self.sh.column_dimensions["I"].width = 10
                self.sh.column_dimensions["J"].width = 20
                self.sh.column_dimensions["K"].width = 20
            # 获取表单的表头信息
            titles = []
            if sheet_name == "支出":
                titles = ["日期", "事项", "支出金额", "支出途径", "备注", "交易方", "一级分类", "二级分类", "对象", "创建时间", "修改时间"]
            elif sheet_name == "收入":
                titles = ["日期", "事项", "收入金额", "收入途径", "备注", "交易方", "一级分类", "二级分类", "对象", "创建时间", "修改时间"]
            elif sheet_name in ["借入", "借出", "还款"]:
                titles = ["日期", "事项", "金额", "账户", "备注", "交易方", "创建时间", "修改时间"]
            else:
                titles = ["日期", "事项", "记录", "额外备注", "创建时间", "修改时间"]
            # print(sheet_name)
            # print(titles)

            # 插入表头
            col = 1
            row = 1
            for title in titles:
                self.sh.cell(row=row, column=col, value=title)
                col += 1
            data = self.result[sheet_name]
            # 插入数据
            for note in data:
                row += 1
                col = 1
                for i in note:
                    # print(i)
                    self.sh.cell(row=row, column=col, value=i)
                    col += 1

        self.wb.remove(self.wb["Sheet"])  # 删除默认创建的空工作簿
        # 保存到文件并关闭工作簿
        self.wb.save(self.file_name)
        self.wb.close()

